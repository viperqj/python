#example7_9.py
#coding=gbk
import openpyxl
#从xlsx文件中导入工作簿对象
workbook = openpyxl.load_workbook('太谷美食商家数据.xlsx')
#从工作簿中获得以sheet名为元素的列表
sheetNames = workbook.sheetnames
#遍历每个sheet
for sheetName in sheetNames:
    print('当前工作表名称为：%s'%sheetName)
    #根据sheet名获取sheet对象
    sheet = workbook[sheetName]
    #遍历sheet中的每个单元格
    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            print(sheet.cell(row=i+1, column=j+1).value,end='\t')
        print()
